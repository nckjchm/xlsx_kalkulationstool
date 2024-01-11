from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.cell import get_column_letter
from kalkulationsdaten import Kalkulationsdaten
import stylingdaten

class Formatierer:
    def __init__(self, daten : Kalkulationsdaten):
        self.daten = daten
        self.daten.filtern()
        self.tabelle : stylingdaten.Tabelle = stylingdaten.Tabelle(daten=self.daten)

    def format_einlesen(self, ws : Worksheet):
        header_ende = self.header_einlesen(ws)
        elemente_ende = self.elemente_einlesen(ws, header_ende + 1)
        self.tabelle.groesse_berechnen()

    def header_einlesen(self, ws : Worksheet):
        if ws['A1'].value != "[HEADER]":
            return 1
        arbeitszeile = 2
        header_daten = []
        while not ist_zeile_ein_blockuebergang(ws, arbeitszeile):
            if not ist_zeile_leer(ws, arbeitszeile):
                header_daten.append(header_zeile_einlesen(ws, arbeitszeile))
            arbeitszeile += 1
        return arbeitszeile

    def elemente_einlesen(self, ws, startzeile):
        if ws['A'+str(startzeile)].value == "[ELEMENTE]":
            startzeile += 1
        while not ist_zeile_ein_blockuebergang(ws, startzeile):
            zeilen_versatz = self.element_einlesen(ws, startzeile)
            startzeile += zeilen_versatz + 1

    def element_einlesen(self, ws, zeile):
        neues_element = stylingdaten.Layout.LayoutElement()
        neues_element.einlesen(ws, zeile)
        self.tabelle.layout.elemente.append(neues_element)
        return neues_element.hoehe
    
    def element_einfuegen(self, ws : Worksheet, start_koordinaten : (int, int), datenpunkt):
        element : stylingdaten.Layout.LayoutElement = self.tabelle.layout.element_suchen(datenpunkt["name"])
        if element is None:
            return None
        element.anwenden(ws, start_koordinaten, datenpunkt["formdaten"])
        return element
    
    def tabelle_einfuegen(self, ws : Worksheet, start_koordinaten : (int, int)):
        ws.delete_rows(start_koordinaten[0])
        ws.insert_rows(start_koordinaten[0], self.tabelle.hoehe)
        arbeitszeile = start_koordinaten[0]
        header_element = self.element_einfuegen(ws, start_koordinaten, {
            "name" : "Header", 
            "formdaten" : self.daten.format_string()})
        arbeitszeile += header_element.hoehe
        for kategorie in self.daten.abrechnungskategorien:
            kategorie_element = self.element_einfuegen(ws, (arbeitszeile, start_koordinaten[1]), {
                "name" : "Kategorie", 
                "formdaten" : self.daten.format_string(kategorie)})
            arbeitszeile += kategorie_element.hoehe
            for einheit in kategorie.rechnungseinheiten:
                einheit_element = self.element_einfuegen(ws, (arbeitszeile, start_koordinaten[1]), {
                    "name" : "Einheit",
                    "formdaten" : self.daten.format_string(kategorie, einheit)
                })
                arbeitszeile += einheit_element.hoehe
        footer_element = self.element_einfuegen(ws, (arbeitszeile, start_koordinaten[1]), {
            "name" : "Footer",
            "formdaten" : self.daten.format_string()
        })


def header_zeile_einlesen(ws : Worksheet, zeile : int):
        daten = []
        for spalte in range(1,10):
            daten.append(ws[get_column_letter(spalte)+str(zeile)])
        return daten

def ist_zeile_leer(ws : Worksheet, zeile : int):
    name = ws['A'+str(zeile)].value
    return name == "" or name is None

def ist_zeile_ein_blockuebergang(ws : Worksheet, zeile : int):
    name = ws['A'+str(zeile)].value
    if name is not None:
        return name[0] == '['
    return False
