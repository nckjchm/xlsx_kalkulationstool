from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from kalkulationsdaten import Kalkulationsdaten
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.cell import get_column_letter
from copy import copy
import referenzstyle

class Stylingdaten:
    def __init__(self):
        self.font : Font = None
        self.fuellung : PatternFill = None
        self.rand : Border = None
        self.anordnung : Alignment = None
        self.zahlenformat : str = None
        self.zellschutz : Protection = None
        self.formatierung : str = None

    #Spaghetti Code, sollte mit einem angebrachten Pattern refaktorisiert werden
    def anwenden(self, zelle : Cell, formdaten : dict):
        if self.font is None:
            zelle.font = referenzstyle.font
        else:
            zelle.font = self.font
        if self.fuellung is None:
            zelle.fill = referenzstyle.fill
        else:
            zelle.fill = self.fuellung
        if self.rand is None:
            zelle.border = referenzstyle.border
        else:
            zelle.border = self.rand
        if self.anordnung is None:
            zelle.alignment = referenzstyle.alignment
        else:
            zelle.alignment = self.anordnung
        if self.zahlenformat is None:
            zelle.number_format = referenzstyle.number_format
        else:
            zelle.number_format = self.zahlenformat
        if self.zellschutz is None:
            zelle.protection = referenzstyle.protection
        else:
            zelle.protection = self.zellschutz
        if self.formatierung is not None:
            if self.formatierung in formdaten:
                zelle.value = formdaten[self.formatierung]
            else:
                zelle.value = self.formatierung
    
    def einlesen(self, zelle : Cell):
        self.font = copy(zelle.font)
        self.fuellung = copy(zelle.fill)
        self.rand = copy(zelle.border)
        self.anordnung = copy(zelle.alignment)
        self.zahlenformat = copy(zelle.number_format)
        self.zellschutz = copy(zelle.protection)
        self.formatierung = copy(zelle.value)

class Tabelle:
    def __init__(self, daten : Kalkulationsdaten = None, layout = None):
        self.daten = daten
        self.layout : Layout = layout
        if self.layout is None:
            self.layout = Layout()
        self.breite : int = 0
        self.hoehe : int = 0
        self.zellen : list[list[Tabelle.Zelldaten]] = None
        self.valide = False

    def groesse_anpassen(self, breite : int, hoehe : int):
        self.breite = breite
        self.hoehe = hoehe
        self.zellen_neugenerieren()

    def zellen_neugenerieren(self):
        self.zellen = [[None for _ in range(self.breite)] for _ in range(self.hoehe)]

    def groesse_berechnen(self):
        if self.layout is None:
            self.valide = False
            return
        if self.daten is None:
            self.valide = False
            return
        self.hoehe = self.layout.zeilen_berechnen(self.daten)
        self.breite = self.layout.breite_bestimmen()
        self.valide = True

    class Zelldaten:
        def __init__(self, wert = None, styling : Stylingdaten = None):
            self.wert = wert
            self.styling = styling

class Layout:
    def __init__(self):
        self.gesamtbreite : int = 0
        self.header_zeilen : int = 1
        self.footer_zeilen : int = 1
        self.elemente : list[Layout.LayoutElement] = []

    def breite_bestimmen(self):
        self.gesamtbreite = max([element.breite for element in self.elemente])

    def zeilen_berechnen(self, daten : Kalkulationsdaten):
        return self.header_zeilen+self.footer_zeilen + len(daten.abrechnungskategorien) + sum([len(kategorie.rechnungseinheiten) for kategorie in daten.abrechnungskategorien])

    def element_suchen(self, name : str):
        for element in self.elemente:
            if element.name == name:
                return element
        print(f"Konnte LayoutElement mit Namen {name} nicht finden")

    class LayoutElement:
        def __init__(self):
            self.breite : int = 0
            self.hoehe : int = 0
            self.name : str = ""
            self.styling : list[list[Stylingdaten]] = None

        def einlesen(self, ws : Worksheet, start_zeile : int):
            self.name = ws['A'+str(start_zeile)].value
            self.breite = ws['B'+str(start_zeile)].value
            self.hoehe = ws['C'+str(start_zeile)].value
            self.styling = [[Stylingdaten() for _ in range(self.breite)] for _ in range(self.hoehe)]
            start_zeile += 1
            for zeile in range(self.hoehe):
                for spalte in range(self.breite):
                    aktuelle_zeile = zeile + start_zeile
                    zellstyling : Stylingdaten = self.styling[zeile][spalte]
                    zellstyling.einlesen(ws[get_column_letter(spalte + 1)+str(aktuelle_zeile)])

        def anwenden(self, ws : Worksheet, start_koordinaten : (int, int), formdaten):
            for zeile in range(self.hoehe):
                arbeitszeile = start_koordinaten[0] + zeile
                for spalte in range(self.breite):
                    arbeitsspalte = start_koordinaten[1] + spalte
                    self.styling[zeile][spalte].anwenden(ws.cell(arbeitszeile, arbeitsspalte), formdaten)
