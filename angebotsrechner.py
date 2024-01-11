from openpyxl import Workbook
from argparse import ArgumentParser
from os.path import abspath
from kalkulationsdaten import Kalkulationsdaten
from util import *
from openpyxl.drawing.image import Image
from openpyxl.utils.cell import get_column_letter
from formatierer import Formatierer
formatvorlage = "resourcen/Kalkulationsvorlage.xlsx"
logodatei = "resourcen/logo.png"
daten_auf_konsole_ausgeben = False

#Anfangspunkt bei direkter Ausführung des Moduls
def main():
    parser = ArgumentParser()
    #Dateinamen einlesen
    parser.add_argument('dateien', nargs='+', help='Liste von Dateinamen.')
    args = parser.parse_args()
    #Kalkulation für jede Input-Datei erstellen
    dateien_abwickeln(args.dateien)
    
def dateien_abwickeln(dateinamen : list[str]):
    for dateiname in dateinamen:
        datei_abwickeln(dateiname)

def datei_abwickeln(dateiname : str):
    dateiname = abspath(dateiname)
    if workbook_pfad_verifizieren(dateiname):
        daten = lese_kalkulationsdaten(dateiname)
        if (daten_auf_konsole_ausgeben):
            daten_auf_konsole_drucken(daten)
        xlsx = vervielfaeltige_formatvorlage(dateiname)
        if xlsx is None:
            return
        wb = workbook_oeffnen(xlsx)
        ws = wb["Kalkulation"]
        schreibe_formatierte_kalkulationsdaten(daten, ws, wb["_styling"])
        wb.save(xlsx)
        fuege_logo_ein(ws)
        wb.save(xlsx)
        wb.close()
    else:
        print(f"Konnte keine Datei unter dem Namen/Pfad {dateiname} finden")


#Steuert die Kalkulationserstellung für eine einzelne Inputdatei
def lese_kalkulationsdaten(dateiname : str):
    datei : Workbook = workbook_oeffnen(dateiname)
    daten : Kalkulationsdaten = daten_einlesen(datei)
    return daten

def vervielfaeltige_formatvorlage(input_pfad : str):
    pfad_formatvorlage = pfad_zu_ressource_finden(formatvorlage)
    pfad_fuer_kalkulationsdatei = neuen_dateipfad_ermitteln(input_pfad)
    kalkulationsdatei = workbook_kopieren(pfad_formatvorlage, pfad_fuer_kalkulationsdatei)
    if kalkulationsdatei is None:
        print("Fehler beim vervielfältigen der Formatvorlage, abbruch.")
        return None
    else:
        return pfad_fuer_kalkulationsdatei
    
def fuege_logo_ein(ws : Worksheet):
    pfad_logo = pfad_zu_ressource_finden(logodatei)
    logoposition = format_logoposition_finden(ws)
    if logoposition is not None:
        bild = Image(pfad_logo)
        logozelle = ws.cell(logoposition[0], logoposition[1])
        logozelle_sting_index = f"{get_column_letter(logoposition[1])}{str(logoposition[0])}"
        logozellen : str = f"{logozelle_sting_index}:{logozelle_sting_index}"
        if ws.merged_cells.__contains__(logozelle_sting_index):
            for bereich in ws.merged_cells:
                if bereich.__contains__(logozelle_sting_index):
                    logozellen_bounds = bereich.bounds
                    logozellen = f"{get_column_letter(logozellen_bounds[1])}{str(logozellen_bounds[0])}:{get_column_letter(logozellen_bounds[3])}{str(logozellen_bounds[2])}"
        else:
            print(f"{logozelle} is not part of any merged cells.")
        gesamtbreite = 0
        for spalte in range(ord(logozellen[0]), ord(logozellen[3]) + 1):
            gesamtbreite += ws.column_dimensions[get_column_letter(spalte)].width
            print(gesamtbreite)
        bild.anchor = logozelle_sting_index
        #bild.width = gesamtbreite
        ws.add_image(bild)

def schreibe_formatierte_kalkulationsdaten(daten : Kalkulationsdaten, ws_output : Worksheet, ws_format : Worksheet):
    startpunkt = format_start_finden(ws_output)
    if startpunkt is not None:
        formatierer = Formatierer(daten)
        formatierer.format_einlesen(ws_format)
        formatierer.tabelle_einfuegen(ws_output, startpunkt)

        """
        formatierte_daten = daten.gefiltert_ausgeben()
        formatierte_daten_einfuegen(ws, formatierte_daten, startpunkt)
        """

#Druckt einen Datensatz Kalkulationsdaten als Baumdiagramm auf die Konsole
def daten_auf_konsole_drucken(daten : Kalkulationsdaten):
    print("Daten erhalten: ", daten)
    for kategorie in daten.abrechnungskategorien:
        print("├─Kategorie:", kategorie.name)
        for einheit in kategorie.rechnungseinheiten:
            print("│  ├─Rechnungseinheit: ", einheit.name, " \n│  │  ├─Preis: ", einheit.preis, " \n│  │  ├─Nenner: ", einheit.nenner)

if __name__ == "__main__":
    daten_auf_konsole_ausgeben = True
    main()