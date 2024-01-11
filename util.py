from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import os
import sys
import shutil
from kalkulationsdaten import Kalkulationsdaten, Rechnungseinheit, Kategorie

def pfad_zu_ressource_finden(resource : str):
    if getattr(sys, 'frozen', False):
        path = os.path.abspath(os.path.join(getattr(sys, '_MEIPASS'), resource))
        return path
    else:
        return resource

def workbook_oeffnen(pfad : str, read_only_modus = False):
    if workbook_pfad_verifizieren(pfad):
        return _workbook_oeffnen(pfad, read_only_modus=read_only_modus)
    
def _workbook_oeffnen(pfad : str, read_only_modus = False):
    return load_workbook(pfad, read_only=read_only_modus)
    
def workbook_kopieren(vorlage : str, neuer_pfad : str):
    if workbook_pfad_verifizieren(vorlage) and validen_neupfad_verifizieren(neuer_pfad):
        return _workbook_kopieren(vorlage, neuer_pfad)

def _workbook_kopieren(vorlage : str, neuer_pfad : str):
    return shutil.copy(vorlage, neuer_pfad)

def workbook_pfad_verifizieren(pfad : str):
    if not os.path.exists(pfad):
        print (f"Datei {pfad} existiert nicht")
        return False
    if not os.path.isfile(pfad):
        print(f"Pfad {pfad} existiert, aber ist keine Datei")
        return False
    if not os.path.splitext(pfad)[-1] == ".xlsx":
        print(f"Datei {pfad} existiert aber ist keine .xlsx Datei")
        return False
    return True

def validen_neupfad_verifizieren(pfad : str):
    if os.path.exists(pfad):
        print("Pfad existiert bereits")
        return False
    if not os.path.isdir(os.path.dirname(pfad)):
        print(f"Überverzeichnis {os.path.dirname(pfad)} existiert nicht")
        return False
    if not os.path.splitext(pfad)[-1] == ".xlsx":
        print("Angegebener Kopiepfad muss eine .xlsx Datei sein")
    #TODO Schreibberechtigung im Verzeichnis abfragen
    return True

def neuen_dateipfad_ermitteln(input_pfad : str, postfix = "Kalkulation"):
    neuer_pfad = os.path.splitext(input_pfad)[0] + "_" + postfix + ".xlsx"
    return neuer_pfad

#Liest die Daten einer Excel Tabelle ein und parset sie in das Kalkulationsdatenschema
def daten_einlesen(datei : Workbook):
    sheet = datei.active
    kalkulationsdaten : Kalkulationsdaten = Kalkulationsdaten()
    letzte_kategorie = ""
    keine_daten_seit = 0
    maximale_sprungweite = 2
    zeile = 1
    while(True):
        gelesene_kategorie = sheet["A" + str(zeile)].value
        rechnungseinheit = sheet["B" + str(zeile)].value
        einheitspreis = sheet["C" + str(zeile)].value
        einheitsnenner = sheet["D" + str(zeile)].value
        einheitsmenge = sheet["E" + str(zeile)].value
        if all(v is None for v in [gelesene_kategorie, rechnungseinheit, einheitspreis, einheitsnenner, einheitsmenge]):
            keine_daten_seit += 1
            if keine_daten_seit > maximale_sprungweite:
                break
            continue
        else:
            keine_daten_seit = 0
        if isinstance(gelesene_kategorie, str):
            gelesene_kategorie = gelesene_kategorie.strip()
        aktuelle_kategorie = gelesene_kategorie
        if gelesene_kategorie == None or gelesene_kategorie == "":
            aktuelle_kategorie = letzte_kategorie
        kategorie_objekt : Kategorie = kalkulationsdaten.kategorie_suchen(aktuelle_kategorie)
        if kategorie_objekt is None:
            kategorie_objekt = Kategorie(aktuelle_kategorie)
            kalkulationsdaten.abrechnungskategorien.append(kategorie_objekt)
        einheit_objekt = Rechnungseinheit(rechnungseinheit, einheitspreis, einheitsmenge, einheitsnenner)
        kategorie_objekt.rechnungseinheiten.append(einheit_objekt)
        letzte_kategorie = aktuelle_kategorie
        zeile += 1
    return kalkulationsdaten

def format_start_finden(ws : Worksheet):
    return formatierungspunkt_finden(ws, "[START]")

def format_logoposition_finden(ws: Worksheet):
    return formatierungspunkt_finden(ws, "[LOGO]")

def formatierungspunkt_finden(ws : Worksheet, muster : str):
    for z_index, zeile in enumerate(ws.iter_rows(max_row = 100, max_col = 100)):
        for s_index, cell in enumerate(zeile):
            if cell.value == muster:
                return z_index + 1, s_index + 1
    return None
